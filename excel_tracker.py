import os
import sys
from openpyxl import Workbook, load_workbook
from datetime import datetime
import time

# Cross-platform file locking
if sys.platform == 'win32':
    import msvcrt
    def lock_file(f):
        try:
            msvcrt.locking(f.fileno(), msvcrt.LK_NBLCK, 1)
            return True
        except OSError:
            return False
    
    def unlock_file(f):
        try:
            msvcrt.locking(f.fileno(), msvcrt.LK_UNLCK, 1)
        except:
            pass
else:
    import fcntl
    def lock_file(f):
        try:
            fcntl.flock(f.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
            return True
        except OSError:
            return False
    
    def unlock_file(f):
        try:
            fcntl.flock(f.fileno(), fcntl.LOCK_UN)
        except:
            pass

EXCEL_FILE = "ticket_records.xlsx"

def get_next_booking_id(platform_code="AT"):
    """Generates the next booking ID automatically by reading the excel file."""
    if not os.path.exists(EXCEL_FILE):
        return f"{platform_code}-{datetime.now().strftime('%Y%m%d')}-0001"
    
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        sno = max(0, ws.max_row - 1)  # Subtract 1 for header row
        return f"{platform_code}-{datetime.now().strftime('%Y%m%d')}-{sno+1:04d}"
    except Exception as e:
        print(f"Error reading Excel for booking ID: {e}")
        return f"{platform_code}-{datetime.now().strftime('%Y%m%d')}-0001"

def save_to_excel(data):
    """Saves ticket data to the excel tracker with proper file locking."""
    # Retry logic for file locking
    max_retries = 3
    retry_count = 0
    
    while retry_count < max_retries:
        try:
            # Attempt to lock file
            with open(EXCEL_FILE, 'a+b') as f:
                locked = lock_file(f)
                if not locked and retry_count < max_retries - 1:
                    retry_count += 1
                    time.sleep(0.5)  # Wait and retry
                    continue
                
                try:
                    # If file exists but is not a valid xlsx (corrupted or wrong format), move it aside and create new
                    if os.path.exists(EXCEL_FILE) and os.path.getsize(EXCEL_FILE) > 0:
                        try:
                            with open(EXCEL_FILE, 'rb') as fh:
                                sig = fh.read(4)
                        except Exception:
                            sig = b''
                        if not sig.startswith(b'PK'):
                            # rename corrupted file
                            corrupt_name = EXCEL_FILE + f'.corrupt.{int(time.time())}'
                            try:
                                os.replace(EXCEL_FILE, corrupt_name)
                                print(f"Renamed corrupted Excel {EXCEL_FILE} to {corrupt_name}")
                            except Exception:
                                # If rename fails, continue and attempt to create a new workbook anyway
                                pass

                    if not os.path.exists(EXCEL_FILE) or os.path.getsize(EXCEL_FILE) == 0:
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Bookings"
                        headers = [
                            "S.No.", "Generated On", "Booking Platform", "PNR", "Booking ID", 
                            "Lead Passenger", "Total Pax", "Customer Phone", "Route", "Travel Date", 
                            "Departure Time", "Flight No(s)", "Total Amount", "Payment Mode", 
                            "Fare Type", "Refund Status", "Flight Status", "Notes"
                        ]
                        ws.append(headers)
                    else:
                        wb = load_workbook(EXCEL_FILE)
                        ws = wb.active

                    # S.No. is the current row count (excluding header)
                    sno = max(1, ws.max_row)
                    data_to_insert = [sno] + data
                    ws.append(data_to_insert)
                    
                    # Auto-adjust column widths
                    for col in ws.columns:
                        max_length = 0
                        column_letter = col[0].column_letter
                        for cell in col:
                            try:
                                cell_str = str(cell.value or "")
                                max_length = max(max_length, len(cell_str))
                            except (TypeError, AttributeError):
                                pass
                        adjusted_width = min(max_length + 2, 50)  # Cap at 50
                        ws.column_dimensions[column_letter].width = adjusted_width

                    wb.save(EXCEL_FILE)
                    return True
                    
                finally:
                    unlock_file(f)
            break
        
        except Exception as e:
            retry_count += 1
            if retry_count >= max_retries:
                print(f"Error saving to Excel after {max_retries} retries: {e}")
                # Fallback: try without locking
                try:
                    if not os.path.exists(EXCEL_FILE):
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Bookings"
                        headers = [
                            "S.No.", "Generated On", "Booking Platform", "PNR", "Booking ID", 
                            "Lead Passenger", "Total Pax", "Customer Phone", "Route", "Travel Date", 
                            "Departure Time", "Flight No(s)", "Total Amount", "Payment Mode", 
                            "Fare Type", "Refund Status", "Flight Status", "Notes"
                        ]
                        ws.append(headers)
                    else:
                        wb = load_workbook(EXCEL_FILE)
                        ws = wb.active

                    sno = max(1, ws.max_row)
                    data_to_insert = [sno] + data
                    ws.append(data_to_insert)
                    wb.save(EXCEL_FILE)
                    return True
                except Exception as e2:
                    print(f"Fallback Excel save also failed: {e2}")
                    return False
            else:
                time.sleep(0.2)
